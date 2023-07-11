from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import check_password
from .models import *
from .forms import CreateUserForm
from django.contrib import messages
from django.views.generic.base import TemplateView
from openpyxl import Workbook #Library for generating an excel doc
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from datetime import date,time,timedelta, datetime

# Create your views here.
from django.http import HttpResponse

#---------------get the role--------------#
def getRole(request):
    actualProfile=Profile.objects.get(user=request.user)
    return actualProfile.rol

def redirectInvalidPage(request,listAvailables):
    actualRole=getRole(request)
    if(actualRole not in listAvailables):
        messages.warning(request,"Ups, creo que no tienes acceso a esa página")
        if(actualRole==0):
            return 'reports'
        elif(actualRole==1):
            return 'see_schedules'
        elif(actualRole==2):
            return 'see_schedules'
        elif(actualRole==3):
            return 'see_schedules'
#-----------------------------------------#
def login_page(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)

        if user is not None:
            profile = Profile.objects.get(user=user.id)
            role = profile.rol
            login(request, user)
            messages.success(request, f'¡Bienvenido {user.username}, nos encanta tenerte de vuelta!')

            if role == 0:
                context = {'role':role}
                return redirect('reports')
            elif role == 1:
                context = {'role':role}
                return redirect('see_schedules')
            elif role == 2:
                context = {'role':role}
                return redirect('see_schedules')
            elif role == 3:
                context = {'role':role}
                return redirect('see_schedules')
            else:
                messages.error(request, 'Ocurrio un error inesperado!')
                return redirect('login')
        else:
            messages.error(request, '¡El usuario y la contraseña no coinciden, por favor vuelva a intentarlo!')
            return redirect('login')
    
    if request.method == 'GET':
        return render(request, template_name='login.html')

def home_manager(request):
    return render(request,template_name="0-manager/home-manager.html")

def home_admin(request):
    return render(request,template_name="1-admin/home-admin.html")

def home_adviser(request):
    return render(request,template_name="2-adviser/home-adviser.html")

def home_operator(request):
    return render(request,template_name="3-operator/home-operator.html")

#---------- auxiliar methods for block schedule -------------------#

@login_required(login_url='')
def block_schedule(request):
    warningPage=redirectInvalidPage(request,[0,1])
    if(warningPage):
        return redirect(warningPage)
    
    if(request.method=='POST'):
        #get the selected name to show the preview
        cvsName=request.POST.get('cvsName')
        #events' array (freetime and services) for the context
        sc_days=toSee_schedules(cvsName)
        #cvs' names for the combo box already selected
        listCVSs=listCVS(cvsName)

        minDate=date.today().isoformat()

        return render(request,template_name="0-1-block_schedule.html", context={'cvsName':cvsName,'sc_days':sc_days,'listCVSs':listCVSs,'minDate':minDate,'role':getRole(request)})
    else:
        if(getRole(request)==1):
            #we take the name of the admin -> cvsName
            cvsName=request.user.first_name
            #events' array (freetime and services) for the context
            sc_days=toSee_schedules(cvsName)
            #cvs' names for the combo box already selected
            listCVSs=listCVS(cvsName)

            minDate=date.today().isoformat()
            
            return render(request,template_name="0-1-block_schedule.html", context={'cvsName':cvsName,'sc_days':sc_days,'listCVSs':listCVSs,'minDate':minDate,'role':getRole(request)})

        else:
            #cvs' names for the combo box
            listCVSs=listCVS('')
            return render(request,template_name="0-1-block_schedule.html", context={'listCVSs':listCVSs,'role':getRole(request)})

@login_required(login_url='')
def block(request):
    warningPage=redirectInvalidPage(request,[0,1])
    if(warningPage):
        return redirect(warningPage)
    
    if(request.method=='POST'):
        cvsName = request.POST.get('cvsName')
        sDate = date.fromisoformat(request.POST.get('startDate'))
        sHour = time.fromisoformat(request.POST.get('startHour'))
        fDate = date.fromisoformat(request.POST.get('finalDate'))
        fHour = time.fromisoformat(request.POST.get('finalHour'))
        _comment = request.POST.get('comment')

        if(fDate.weekday()==5):
            fHour = time.fromisoformat('12:00')

        #if is only one day then is from the start hour to the final hour
        if(sDate == fDate):
            beg = timedelta(hours=sHour.hour, minutes=sHour.minute)
            end = timedelta(hours = fHour.hour, minutes=fHour.minute)
            difference =  end - beg
            differenceMinutes = difference.total_seconds()/60
            duration=int(differenceMinutes)
            flag=valDateHour(request,cvsName,sDate,sHour,duration,1,None,True)
            flag1 = valBlocks(request, cvsName, sDate, duration, sHour)
            if(flag or flag1):
                messages.error(request,"Hay turnos asignados para esta fecha, por favor modifiquelos antes de seguir con el bloqueo.")
            else:
                newBlock=Block(
                    cvs = CVS.objects.get(name=cvsName),
                    duration = duration,
                    startDate = sDate,
                    startHour = sHour,
                    endDate = fDate,
                    endHour = fHour,
                    comment = _comment,
                    scheduledBy = Profile.objects.get(user=request.user)
                )
                newBlock.save()
                messages.success(request,"Se ha agendado correctamente el bloqueo en CVS: "+cvsName+" el día "+sDate.strftime("%d/%m/%Y")+" a las "+sHour.strftime("%I:%M %p"))
                
                return redirect('see_schedules')
        #Check if is more than one day, if so from the first date we have to do a for, from the first hour until 18pm
        #in the first day, then the other is from 7am to 18pm, then when we get to the last day we do it from 7am
        #until the final hour.
        else:
            day = sDate
            totalDays = fDate-sDate
            totalDays = int((totalDays.total_seconds()/3600)/24)
            for i in range (0,totalDays+1):
                if (i == 0):
                    beg = timedelta(hours=sHour.hour, minutes=sHour.minute)
                    if(day.weekday()==5):
                        end = timedelta(hours = 12, minutes=00)
                    else:
                        end = timedelta(hours = 17, minutes=15)
                    difference =  end - beg
                    differenceMinutes = difference.total_seconds()/60
                    duration=int(differenceMinutes)
                    flag=valDateHour(request,cvsName,day,sHour,duration,1,None,True)
                    flag1 = valBlocks(request, cvsName, day, duration, sHour)
                    if(flag or flag1):
                        messages.error(request,"Hay turnos asignados para esta fecha, por favor modifiquelos antes de seguir con el bloqueo.")
                        break
                    else:
                        if(day.weekday()==5):
                            end = time.fromisoformat('12:00')
                        else:
                            end = time.fromisoformat('17:15')
                        newBlock=Block(
                            cvs = CVS.objects.get(name=cvsName),
                            duration = duration,
                            startDate = day,
                            startHour = sHour,
                            endDate = day,
                            endHour = end,
                            comment = _comment,
                            scheduledBy = Profile.objects.get(user=request.user)
                        )
                        newBlock.save()
                        messages.success(request,"Se ha agendado correctamente el bloqueo en CVS: "+cvsName+" el día "+sDate.strftime("%d/%m/%Y")+" a las "+sHour.strftime("%I:%M %p"))

                elif (i == totalDays):
                    beg=timedelta(hours=7, minutes=45)
                    end = timedelta(hours = fHour.hour, minutes=fHour.minute)
                    difference =  end - beg
                    differenceMinutes = difference.total_seconds()/60
                    duration=int(differenceMinutes)
                    flag=valDateHour(request,cvsName,day,time.fromisoformat('07:45'),duration,1,None,True)
                    flag1 = valBlocks(request, cvsName, day, duration, time.fromisoformat('07:45'))
                    if(flag or flag1):
                        messages.error(request,"Hay turnos asignados para esta fecha, por favor modifiquelos antes de seguir con el bloqueo.")
                        break
                    else:
                        newBlock=Block(
                            cvs = CVS.objects.get(name=cvsName),
                            duration = duration,
                            startDate = day,
                            startHour = time.fromisoformat('07:45'),
                            endDate = day,
                            endHour = fHour,
                            comment = _comment,
                            scheduledBy = Profile.objects.get(user=request.user)
                        )
                        newBlock.save()
                        messages.success(request,"Se ha agendado correctamente el bloqueo en CVS: "+cvsName+" el día "+day.strftime("%d/%m/%Y")+" a las 7:45am")

                else:
                    beg=timedelta(hours=7, minutes=45)
                    if(day.weekday()==5):
                        end = timedelta(hours = 12, minutes=00)
                    else:
                        end = timedelta(hours = 17, minutes=15)
                    difference =  end - beg
                    differenceMinutes = difference.total_seconds()/60
                    duration=int(differenceMinutes)
                    flag=valDateHour(request,cvsName,day,time.fromisoformat('07:45'),duration,1,None,True)
                    flag1 = valBlocks(request, cvsName, day, duration, time.fromisoformat('07:45'))
                    if(flag or flag1):
                        messages.error(request,"Hay turnos asignados para esta fecha, por favor modifiquelos antes de seguir con el bloqueo.")
                        break
                    else:
                        if(day.weekday()==5):
                            end = time.fromisoformat('12:00')
                        else:
                            end = time.fromisoformat('17:15')
                        
                        newBlock=Block(
                            cvs = CVS.objects.get(name=cvsName),
                            duration = duration,
                            startDate = day,
                            startHour = time.fromisoformat('07:45'),
                            endDate = day,
                            endHour = end,
                            comment = _comment,
                            scheduledBy = Profile.objects.get(user=request.user)
                        )
                        newBlock.save()
                        messages.success(request,"Se ha agendado correctamente el bloqueo en CVS: "+cvsName+" el día "+day.strftime("%d/%m/%Y")+" a las 7:45am")

                day = day + timedelta(days=1)
                print(day)
            return redirect('see_schedules')

        
        #then we have to see what we are going to do with the turns assigned in this dates and hours

        

    return redirect('see_schedules')

#---------- methods for users -------------------#

@login_required(login_url='')
def create_user(request):
    form = CreateUserForm()

    if request.method == "POST":
        form = CreateUserForm(request.POST)
        if form.is_valid():
            role = request.POST.get('rol') #Revisar que el name sea rol
            form.save()
            new_user = User.objects.get(username= request.POST["username"])
            newProfile = Profile(user=new_user, rol=role)
            newProfile.save()
            messages.success(request, '¡El usuario fue creado exitosamente!')
            return render(request,template_name="0-1-create_user.html", context={'form': form, 'role':getRole(request)})
        else:
            messages.error(request, 'Error al crear el usuario, la contraseña debe tener minimo 8 caracteres, una mayuscula, una minuscula y un simbolo ej(#$%&) o esta usuario ya existe')
            return render(request,template_name="0-1-create_user.html", context={'form': form, 'role':getRole(request)})

    return render(request,template_name="0-1-create_user.html", context={'form': form, 'role':getRole(request)})

def change_password(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('current_password')
        user = authenticate(username=username, password=password)

        if user is not None:
            new_password1 = request.POST.get('new_password')
            new_password2 = request.POST.get('confirmation_password')
            if(new_password1 == new_password2):
                user.set_password(new_password1)
                user.save()
                user.save()
                messages.success(request, '¡La contraseña fue actualizada exitosamente!')
                return redirect('login')
            else:
                messages.success(request, '¡La nueva contraseña no coincide con la de confirmación, vuelva a intentarlo!')
                return render(request,template_name="change_password.html")
        else:
            messages.success(request, '¡El usuario ingresado no existe!')
            return render(request,template_name="change_password.html")
    return render(request,template_name="change_password.html")

@login_required(login_url='')
def delete_user(request):    

    if request.method == 'POST':
        user_deleted = request.POST.get('user_deleted')
        user_to_delete = User.objects.get(id=user_deleted)
        profile_to_delete = Profile.objects.get(user=user_to_delete)
        profile_to_delete.delete()
        user_to_delete.delete()
    
    current_user = request.user
    users = User.objects.all().order_by('username').values().exclude(id=current_user.id)
    print(users)
    context={'users':users,'role':getRole(request)}
    return render(request,template_name="0-1-delete_user.html",context=context)

#---------- methods for excel report -------------------#

@login_required(login_url='')    
def createReport(request):
    warningPage=redirectInvalidPage(request,[0])
    if(warningPage):
        return redirect(warningPage)
    
    if request.method == 'POST':
        cvsList = request.POST.get('cvsList')
        cvsList = cvsList[1:-1]
        cvsList = cvsList.replace("'","")
        cvsList = cvsList.split(", ")
        initialDate = date.fromisoformat(request.POST.get('initialDate'))
        finalDate = date.fromisoformat(request.POST.get('finalDate'))

        query = Turn.objects.filter(cvs__name__in=cvsList).filter(date__gte=initialDate).filter(date__lte=finalDate).order_by('cvs','date','hour')
        
    else:
        messages.error(request,"No puedes acceder a este apartado sin haber llenado la información del reporte")
        return redirect('reports')   


    workBook = Workbook()
    workSheet = workBook.active
    workSheet.title = 'Hoja1' #Change name of first sheet

    
    #Title of the doc
    workSheet['B2'] = 'REPORTE MONTAJE Y BALANCEO'
    workSheet['B2'].font = Font(name='Arial', size=12, bold=True, color='F87910')
    workSheet['B2'].alignment = Alignment(horizontal="left", vertical="center")
    workSheet.merge_cells('B2:D2')

    workSheet['B3'] = 'Kairos Logistics'
    workSheet['B3'].font = Font(name='Arial', size=12, bold=True, color='CC6600')
    workSheet['B3'].alignment = Alignment(horizontal="left", vertical="center")
    workSheet.merge_cells('B3:D3')


    #Create first row of the table with the data
    workSheet['B5'] = 'Fecha del servicio'
    workSheet['B5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['B5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['B5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['B5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['B'].width = 19

    workSheet['C5'] = 'Hora del servicio'
    workSheet['C5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['C5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['C5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['C5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['C'].width = 18

    workSheet['D5'] = 'CVS'
    workSheet['D5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['D5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['D5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['D5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['D'].width = 17

    workSheet['E5'] = 'Número de factura'
    workSheet['E5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['E5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['E5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['E5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['E'].width = 19.30

    workSheet['F5'] = 'Cédula del cliente'
    workSheet['F5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['F5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['F5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['F5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['F'].width = 19

    workSheet['G5'] = 'Nombre del cliente'
    workSheet['G5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['G5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['G5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['G5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['G'].width = 19.73

    workSheet['H5'] = 'Teléfono del cliente'
    workSheet['H5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['H5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['H5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['H5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['H'].width = 20.58

    workSheet['I5'] = 'Servicio prestado'
    workSheet['I5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['I5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['I5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['I5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['I'].width = 18.30

    workSheet['J5'] = 'Comentarios del servicio'
    workSheet['J5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['J5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['J5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['J5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['J'].width = 25

    workSheet['K5'] = 'Tipo de llanta'
    workSheet['K5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['K5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['K5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['K5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['K'].width = 15

    workSheet['L5'] = 'Cantidad de llantas'
    workSheet['L5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['L5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['L5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['L5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['L'].width = 20

    workSheet['M5'] = '¿Hubo rotación?'
    workSheet['M5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['M5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['M5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['M5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['M'].width = 17.45

    workSheet['N5'] = 'Cantidad de llantas rotadas'
    workSheet['N5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['N5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['N5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['N5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['N'].width = 27.45

    workSheet['O5'] = 'Duración del servicio'
    workSheet['O5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['O5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['O5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['O5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['O'].width = 21.60

    workSheet['P5'] = 'Agendado por'
    workSheet['P5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['P5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['P5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['P5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['P'].width = 16

    workSheet['Q5'] = 'Fecha agendamiento'
    workSheet['Q5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['Q5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['Q5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['Q5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['Q'].width = 21.72

    workSheet['R5'] = 'Modificado por'
    workSheet['R5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['R5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['R5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['R5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['R'].width = 16

    workSheet['S5'] = 'Fecha de modificación'
    workSheet['S5'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    workSheet['S5'].alignment = Alignment(horizontal="center", vertical="center")
    workSheet['S5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    workSheet['S5'].fill = PatternFill(start_color='F87910', end_color='F87910', fill_type="solid")
    workSheet.column_dimensions['S'].width = 23.30

    rowController = 6 #to keep track of the row we will write on
    for q in query:
        #Fill in the data of the table
        workSheet.cell(row=rowController, column=2).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=2).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=2).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        workSheet.cell(row=rowController, column=2).value = q.date

        workSheet.cell(row=rowController, column=3).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=3).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=3).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        workSheet.cell(row=rowController, column=3).value = q.hour

        workSheet.cell(row=rowController, column=4).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=4).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=4).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        workSheet.cell(row=rowController, column=4).value = q.cvs.name

        workSheet.cell(row=rowController, column=5).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=5).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=5).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        workSheet.cell(row=rowController, column=5).value = q.bill

        workSheet.cell(row=rowController, column=6).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=6).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=6).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        workSheet.cell(row=rowController, column=6).value = q.idCustomer

        workSheet.cell(row=rowController, column=7).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=7).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=7).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        workSheet.cell(row=rowController, column=7).value = q.nameCustomer

        workSheet.cell(row=rowController, column=8).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=8).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=8).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        workSheet.cell(row=rowController, column=8).value = q.telCustomer

        workSheet.cell(row=rowController, column=9).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=9).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=9).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        if(q.done is True):
            workSheet.cell(row=rowController, column=9).value = "Si"
        elif(q.done is False):
            workSheet.cell(row=rowController, column=9).value = "No"
        else:
            workSheet.cell(row=rowController, column=9).value = "-"

        workSheet.cell(row=rowController, column=10).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=10).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=10).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        if(q.comment):
            workSheet.cell(row=rowController, column=10).value = q.comment
        else:
            workSheet.cell(row=rowController, column=10).value = "-"

        workSheet.cell(row=rowController, column=11).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=11).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=11).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        if(q.typeTire == 1):
            workSheet.cell(row=rowController, column=11).value = "Automóvil"
        elif(q.typeTire == 2):
            workSheet.cell(row=rowController, column=11).value = "Camioneta"
        elif(q.typeTire == 3):
            workSheet.cell(row=rowController, column=11).value = "Camión"

        workSheet.cell(row=rowController, column=12).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=12).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=12).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        workSheet.cell(row=rowController, column=12).value = q.quantity

        workSheet.cell(row=rowController, column=13).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=13).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=13).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        if(q.rotation is True):
            workSheet.cell(row=rowController, column=13).value = "Si"
        else:
            workSheet.cell(row=rowController, column=13).value = "No"

        workSheet.cell(row=rowController, column=14).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=14).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=14).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        workSheet.cell(row=rowController, column=14).value = q.quantityRotate

        workSheet.cell(row=rowController, column=15).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=15).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=15).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        workSheet.cell(row=rowController, column=15).value = str(q.duration) + " min."

        workSheet.cell(row=rowController, column=16).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=16).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=16).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        workSheet.cell(row=rowController, column=16).value = q.scheduledBy.user.username

        workSheet.cell(row=rowController, column=17).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=17).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=17).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        workSheet.cell(row=rowController, column=17).value = str(q.dateScheduled.date()) + ", " + str(q.dateScheduled.time().isoformat('minutes'))

        workSheet.cell(row=rowController, column=18).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=18).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=18).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        if(q.modifiedBy is None):
            workSheet.cell(row=rowController, column=18).value = ""
        else:
            workSheet.cell(row=rowController, column=18).value = q.modifiedBy.user.username

        workSheet.cell(row=rowController, column=19).alignment = Alignment(horizontal="center", vertical="center")
        workSheet.cell(row=rowController, column=19).font = Font(name='Arial', size=10)
        workSheet.cell(row=rowController, column=19).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        if(q.dateModified is None):
            workSheet.cell(row=rowController, column=19).value = ""
        else:
            workSheet.cell(row=rowController, column=19).value = str(q.dateModified.date()) + ", " + str(q.dateModified.time().isoformat('minutes'))

        rowController += 1

    name = "Reporte.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    content = "attachment; filename = {0}".format(name)
    response["Content-Disposition"] = content
    workBook.save(response)
    return response
   
@login_required(login_url='')
def reports(request):    
    warningPage=redirectInvalidPage(request,[0])
    if(warningPage):
        return redirect(warningPage)

    allCvs = CVS.objects.all()

    if request.method == 'POST':
        cvsList = request.POST.getlist('cvsOptions')
        initialDate = date.fromisoformat(request.POST.get('dateS'))
        finalDate = date.fromisoformat(request.POST.get('dateE'))

        if not(initialDate and finalDate and cvsList):
            messages.warning(request,"Recuerda que debes llenar todos los campos del formulario para el reporte")
            redirect('reports')
            preview = False

        preview = True
        query = Turn.objects.filter(cvs__name__in=cvsList).filter(date__gte=initialDate).filter(date__lte=finalDate).order_by('cvs','date','hour')

        return render(request,template_name="0-reports.html", context={'role':getRole(request), 'allCvs':allCvs, 'preview':preview, 'cvsList':cvsList, 'initialDate':initialDate.isoformat(), 'finalDate':finalDate.isoformat(),'query':query})
    
    
    return render(request,template_name="0-reports.html", context={'role':getRole(request), 'allCvs':allCvs})

#---------- auxiliar methods for see_schedules-------------------#
def delta2time(delta):
    noon=timedelta(hours=12)
    flagNoon=True
    if(delta>=noon):
        delta-=noon
        flagNoon=False

    hours=delta.seconds//3600
    minutes=delta.seconds//60-60*(delta.seconds//3600)

    if(hours==0):
        hours=12

    if(minutes//10>0):
        hour=str(hours)+':'+str(minutes)
    else:
        hour=str(hours)+':0'+str(minutes)

    if(flagNoon):
        hour+=" AM"
    else:
        hour+=" PM"

    return hour

def deltas2line(delta1,delta2):
    li=(delta1.seconds-27000)//60+61
    ls=(delta2.seconds-27000)//60+61
    return(str(li)+"/"+str(ls))

def debugEvent(isFree,delta1,delta2):
    if(isFree):
        print("está libre de "+delta2time(delta1)+" a "+delta2time(delta2))
    else:
        print("está en servicio de "+delta2time(delta1)+" a "+delta2time(delta2))

def fillFree(li,ls,list):
    flagLn=True
    while(li<ls):
        secs=3600-li.seconds%3600

        if(secs==0):
            secs=3600

        if(secs==3600 and (ls-li).seconds < 3600 ):
            secs=(ls-li).seconds

        if(li.seconds+secs>ls.seconds):
            secs=(ls-li).seconds

        increment=timedelta(seconds=secs)

        debugEvent(True,li,li+increment)

        if(flagLn):
            newSC_event=["sc-event free",deltas2line(li,li+increment)]
            flagLn=False
        else:
            newSC_event=["sc-event free ln",deltas2line(li,li+increment)]

        list.append(newSC_event)

        li+=increment

def processData(event):
    info=[]
    #we get the id
    info.append("turn"+str(event.id))
    #we get date
    months=["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    days=["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado"]
    date = event.date

    strDate=days[date.weekday()]+", "+str(date.day)+" de "+months[date.month-1]+", "+str(date.year)
    info.append(strDate)
    #we get cvs
    info.append(event.cvs.name)
    #we get hour
    hour=event.hour
    info.append(hour.strftime("%I:%M %p"))
    #we get duration
    duration=str(event.duration)+" minutos"
    info.append(duration)
    #we get typeTire
    typeTire=event.typeTire
    if(typeTire==1):
        typeTire="Automóvil"
    elif(typeTire==2):
        typeTire="Camioneta"
    elif(typeTire==3):
        typeTire="Camión"
    info.append(typeTire)
    #we get quantity
    quantity=str(event.quantity)+" llantas"
    info.append(quantity)
    #we get rotation
    rotation=event.rotation
    if(rotation):
        rotation="Sí"
    else:
        rotation="No"
    info.append(rotation)
    #we get quantityRotate
    quantityRotate=str(event.quantityRotate)+" llantas"
    info.append(quantityRotate)
    #we get bill
    bill=str(event.bill)
    info.append(bill)
    #we get idCustomer
    idCustomer=str(event.idCustomer)
    info.append(idCustomer)
    #we get nameCustomer
    info.append(event.nameCustomer)
    #we get telCustomer
    info.append(event.telCustomer)
    #we get scheduledBy
    scheduledBy=event.scheduledBy.user.username
    info.append(scheduledBy)
    #we get dateScheduled
    dateScheduled=event.dateScheduled
    info.append(dateScheduled.strftime("%d/%m/%Y a las %I:%M %p"))
    #we get modifiedBy
    if(event.modifiedBy):#revisar que haya modificado
        modifiedBy=event.modifiedBy.user.username
        info.append(modifiedBy)
    else:
        info.append(" ")
    #we get dateModified
    dateModified=event.dateModified
    if(dateModified):#revisar que haya modificado
        info.append(dateModified.strftime("%d/%m/%Y a las %I:%M %p"))
    else:
        info.append(" ")

    return info

def toSee_schedules(cvsName):
    #events' array (freetime and services) for the context
    sc_days=[[],[],[],[],[],[],[]]
    #we start looking for the day of the week
    weekdays=["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado"]
    #the first date 
    actualDate=date.today()
    contDay=0
    while(contDay<7):
        #have to avoid the sunday
        if(actualDate.weekday()==6):
            increment=timedelta(days=1)
            actualDate+=increment
            continue
        
        #name of de day
        day=weekdays[actualDate.weekday()]
        subtitle=["sc-subtitle","1/31",day+' '+str(actualDate.day)+'/'+str(actualDate.month)]
        sc_days[contDay].append(subtitle)
        #change the style of the previous hours
        li=timedelta(hours=7)
        #every day starts at 07:45am
        ti=timedelta(hours=7,minutes=45)
        newSC_event=["sc-event block",deltas2line(li,ti)]
        sc_days[contDay].append(newSC_event)

        #we get actualDate's agenda
        turns = list(Turn.objects.filter(date=actualDate).filter(cvs__name=cvsName).order_by("hour"))
        blocks = list(Block.objects.filter(startDate=actualDate).filter(cvs__name=cvsName).order_by("startHour"))
        agenda=[]
        index = len(turns) + len(blocks)
        i = 0
        while (i<index):
            if(len(turns)):
                if(len(blocks)):
                    turnHour = timedelta(hours=turns[0].hour.hour, minutes=turns[0].hour.minute)
                    blockHour = timedelta(hours=blocks[0].startHour.hour, minutes=blocks[0].startHour.minute)
                    if( turnHour > blockHour):
                        agenda.append(blocks.pop(0))  
                    else:
                        agenda.append(turns.pop(0))
                else:
                    agenda.append(turns.pop(0))
            else:
                agenda.append(blocks.pop(0))
            i+=1

        for event in agenda:
            if(type(event) == type(Turn())):
                beg=timedelta(hours=event.hour.hour,minutes=event.hour.minute)
                fillFree(ti,beg,sc_days[contDay])

                #finally we add the service
                finish=timedelta(minutes=event.duration)
                finish+=beg
                debugEvent(False,beg,finish)

                newSC_event=["sc-event turn",deltas2line(beg,finish),"",delta2time(beg)+" - "+delta2time(finish), processData(event)]
                sc_days[contDay].append(newSC_event)

                ti=finish
            else:
                beg=timedelta(hours=event.startHour.hour,minutes=event.startHour.minute)
                fillFree(ti,beg,sc_days[contDay])

                #finally we add the service
                finish=timedelta(minutes=event.duration)
                finish+=beg
                debugEvent(False,beg,finish)

                newSC_event=["sc-event block",deltas2line(beg,finish)]
                sc_days[contDay].append(newSC_event)

                ti=finish


        #day ends at...
        # if sturday at 12:00m 
        if(actualDate.weekday()==5):
            end=timedelta(hours=12)
        # else at 05:15pm
        else:
            end=timedelta(hours=17,minutes=15)

        if(ti<end):
            fillFree(ti,end,sc_days[contDay])

        #change the style of the last hours
        ls=timedelta(hours=18)
        newSC_event=["sc-event block",deltas2line(end,ls)]
        if(contDay==6):
            newSC_event[1]= str(newSC_event[1])+"; border-bottom-right-radius: 2em"
        sc_days[contDay].append(newSC_event)

        contDay+=1
        actualDate+=timedelta(days=1)
        print("---------end day------------")

    return sc_days

def listCVS(cvsName):
    cvsQuery= CVS.objects.all().order_by("name")

    listNames=[]

    for cvs in cvsQuery:
        listNames.append(cvs.name)

    if(cvsName):
        listNames.remove(cvsName)

    return listNames

@login_required(login_url='')
def see_schedules(request):
    warningPage=redirectInvalidPage(request,[1,2,3])
    if(warningPage):
        return redirect(warningPage)
    
    role=getRole(request)
    print(role)

    if(role==3):
        #we take the name of the admin -> cvsName
        cvsName=request.user.first_name
        #events' array (freetime and services) for the context
        sc_days=toSee_schedules(cvsName)
        return render(request,template_name="1-2-3-see_schedules.html", context={'cvsName':cvsName,'sc_days':sc_days,'role':role})

    
    if(request.method=='POST'):
        cvsName=request.POST.get('cvsName')
        #events' array (freetime and services) for the context
        sc_days=toSee_schedules(cvsName)
        #cvs' names for the combo box
        listCVSs=listCVS(cvsName)
        return render(request,template_name="1-2-3-see_schedules.html", context={'cvsName':cvsName,'sc_days':sc_days,'listCVSs':listCVSs,'role':role})
    else:
        #cvs' names for the combo box
        listCVSs=listCVS('')
        return render(request,template_name="1-2-3-see_schedules.html", context={'listCVSs':listCVSs,'role':role})
    
@login_required(login_url='')
def asign_turns(request):
    warningPage=redirectInvalidPage(request,[1,2])
    if(warningPage):
        return redirect(warningPage)
    
    if(request.method=='POST'):
        #get the selected name to show the preview
        cvsName=request.POST.get('cvsName')
        #events' array (freetime and services) for the context
        sc_days=toSee_schedules(cvsName)
        #cvs' names for the combo box already selected
        listCVSs=listCVS(cvsName)
        return render(request,template_name="1-2-asign_turns.html", context={'cvsName':cvsName,'sc_days':sc_days,'listCVSs':listCVSs,'role':getRole(request)})
    else:
        #cvs' names for the combo box
        listCVSs=listCVS('')
        return render(request,template_name="1-2-asign_turns.html",context={'listCVSs':listCVSs,'role':getRole(request)})

#---------- auxiliar methods for time_turns-------------------#
@login_required(login_url='')
def valQuantity(request,quantity):
    problems=False
    #quantity must be an integer
    if(quantity%1>0):
        problems=True
        messages.error(request,"La cantidad de llantas del servicio no puede ser fraccionaria")
    #quantity must be a positive integer
    if(quantity<=0):
        problems=True
        messages.error(request,"La cantidad de llantas del servicio debe ser mayor que 0")

    return problems

@login_required(login_url='')
def valQuantityRotate(request,quantityRotate,quantity):
    problems=False
    #quantity must be an integer
    if(quantityRotate%1>0):
        problems=True
        messages.error(request,"La cantidad de llantas a rotar no puede ser fraccionaria")
    #quantity must be a positive integer
    if(quantityRotate<=0):
        problems=True
        messages.error(request,"La cantidad de llantas a rotar debe ser mayor que 0")
    #politics says, rotating wheels must be lower or equal than service wheels
    if(quantityRotate>quantity):
        problems=True
        messages.error(request,"La cantidad de llantas a rotar debe ser menor o igual a las llantas del servicio")

    return problems

def calculateTime(typeTire,quantity,rotation,quantityRotate):
    timeService=45
    #how to calculate

    #always add 5 minutes for the rest after service
    timeService+=5

    return timeService

@login_required(login_url='')
def valDuration(request,duration):
    problems=False
    #quantity must be an integer
    if(duration%1>0):
        problems=True
        messages.error(request,"La duración no puede ser fraccionaria")
    #quantity must be a positive integer
    if(duration<=0):
        problems=True
        messages.error(request,"La duración debe ser mayor que 0")

    return problems

def recommendTurns(cvs,duration):
    weekdays=["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado"]
    actualDate=date.today()
    contDay=0
    listOptions=[]
    while(contDay<7):
        weekday=actualDate.weekday()
        #have to avoid the sunday
        if(weekday==6):
            increment=timedelta(days=1)
            actualDate+=increment
            continue

        dateTurns=Turn.objects.filter(cvs__name=cvs).filter(date=actualDate).order_by('hour')

        if (not contDay):
            begHour=datetime.today()
            begHour=timedelta(hours=begHour.hour,minutes=begHour.minute)
        else:
            begHour=timedelta(hours=7,minutes=45)

        for turn in dateTurns:
            toCheck=begHour+timedelta(minutes=duration)
            li=timedelta(hours=turn.hour.hour,minutes=turn.hour.minute)
            if(li>=toCheck):
                break
            else:
                begHour=li+timedelta(minutes=turn.duration)

        endHour=begHour+timedelta(minutes=duration)

        if((weekday!=5 and endHour<=timedelta(hours=17,minutes=15)) or (weekday==5 and endHour<=timedelta(hours=12,minutes=0))):
            listOptions.append(weekdays[weekday]+' '+str(actualDate.day)+' '+delta2time(begHour))        

        actualDate+=timedelta(days=1)
        contDay+=1

    if(not listOptions):
        listOptions.append("No hay horarios disponibles esta semana")
    
    return listOptions

@login_required(login_url='')
def time_turns(request):
    warningPage=redirectInvalidPage(request,[1,2])
    if(warningPage):
        return redirect(warningPage)
    
    if(request.method=='POST'):
        #data to validate
        cvsName=request.POST.get('cvsName')
        typeTire=request.POST.get('typeTire')
        quantity=int(request.POST.get('quantity'))
        flag1=valQuantity(request,quantity)
        if(request.POST.get('rotation')):
            rotation="True"
            quantityRotate=int(request.POST.get('quantityRotate'))
            flag2=valQuantityRotate(request,quantityRotate,quantity)
        else:
            rotation=""
            quantityRotate=0
            flag2=False

        #events' array (freetime and services) for the context
        sc_days=toSee_schedules(cvsName)
        #cvs' names for the combo box
        listCVSs=listCVS(cvsName)
        
        #debug
        print("valores:",cvsName,"|",typeTire,"|",quantity,"|",rotation,"|",quantityRotate)

        #validation
        if(flag1 or flag2):
            return redirect('asign_turns')
        else:
            duration=calculateTime(typeTire,quantity,rotation,quantityRotate)
            recommendations=recommendTurns(cvsName,duration)

            return render(request,template_name="1-2-time_turns.html", context={'cvsName':cvsName,'sc_days':sc_days,'listCVSs':listCVSs,'typeTire':typeTire,'quantity':quantity,'rotation':rotation,'quantityRotate':quantityRotate,'duration':duration,'recommendations':recommendations,'role':getRole(request)})
    else:
        messages.error(request,"No puedes acceder a este apartado sin haber llenado la información inicial del servicio")
        return redirect('asign_turns')   
    
@login_required(login_url='')
def select_turns(request):
    warningPage=redirectInvalidPage(request,[1,2])
    if(warningPage):
        return redirect(warningPage)
    
    if(request.method=='POST'):
        #validated data
        cvsName=request.POST.get('cvsName')
        typeTire=request.POST.get('typeTire')
        quantity=request.POST.get('quantity')
        rotation=request.POST.get('rotation')
        quantityRotate=request.POST.get('quantityRotate')
        duration=request.POST.get('duration')
        #limits for the inputs
        minDate=date.today().isoformat()
        if(date.today().weekday==6):
            maxDate=(date.today()+timedelta(days=7)).isoformat()
        else:
            maxDate=(date.today()+timedelta(days=7)).isoformat()
        deltaMax=timedelta(hours=17,minutes=15)
        deltaDuration=timedelta(minutes=int(duration))
        s = (deltaMax-deltaDuration).seconds
        hours,remainder = divmod(s, 3600)
        minutes= remainder//60
        maxTime=('{:02}:{:02}'.format(int(hours), int(minutes)))
        #debug
        print(minDate)
        print(maxDate)
        print(maxTime)
        #events' array (freetime and services) for the context
        sc_days=toSee_schedules(cvsName)
        #show the recommendations
        recommendations=recommendTurns(cvsName,int(duration))
        return render(request,template_name="1-2-select_turns.html", context={'cvsName':cvsName,'sc_days':sc_days,'typeTire':typeTire,'quantity':quantity,'rotation':rotation,'quantityRotate':quantityRotate,'duration':duration,'minDate':minDate,'maxDate':maxDate,'maxTime':maxTime,'recommendations':recommendations,'role':getRole(request)})
    else:
        messages.error(request,"No puedes acceder a este apartado sin haber llenado la información inicial del servicio")
        return redirect('asign_turns')
    
#---------- auxiliar methods for confirm asignment-------------------#
@login_required(login_url='')
def valDateHour(request,cvs,dateF,hour,duration,modAs,exclude, block):#mod=0,As=1
    problems=False
    if(modAs):
        dateTurns=Turn.objects.filter(cvs__name=cvs).filter(date=dateF).order_by('hour')
        #turns must be scheduled since "today" until 7 valid days(no before)
        if(dateF<date.today()):
            problems=True
            messages.error(request,"El servicio no puede asignarse un día previo a hoy")
    else:
        dateTurns=Turn.objects.filter(cvs__name=cvs).filter(date=dateF).exclude(hour=exclude).order_by('hour')
        print("we are avoiding: ",exclude)

    #turns must be scheduled since "today" until 7 valid days(no after)
    if(date.today().weekday==6):
        #if we look the calendar on sunday there will be 1 day more
        if((dateF-date.today()).days>8):
            problems=True
            messages.error(request,"El servicio no puede asignarse un día posterior a 7 días hábiles")
    else:
        if((dateF-date.today()).days>7):
            problems=True
            messages.error(request,"El servicio no puede asignarse un día posterior a 7 días hábiles")
    actualTurns=[]
    for turn in dateTurns:
        print("actual hour: ",turn.hour)
        beg=timedelta(hours=turn.hour.hour,minutes=turn.hour.minute)
        end=beg+timedelta(minutes=turn.duration)
        actualTurns.append([beg,end])

    begHour=timedelta(hours=hour.hour,minutes=hour.minute)
    endHour=begHour+timedelta(minutes=duration)

    #turns must be scheduled after now (just for assignment)
    nowDate=datetime.today()
    if(modAs and dateF==date.today() and begHour<timedelta(hours=nowDate.hour,minutes=nowDate.minute)):
        problems=True
        messages.error(request,"El servicio no puede asignarse a una hora ya pasada")
    if not(block):
        #turn must not be earlier than 7:45 nor later than 17:15
        if(begHour<timedelta(hours=7,minutes=45)):
            problems=True
            messages.error(request,"El servicio no puede comenzar antes del horario laboral (7:45am)")
        if(endHour>timedelta(hours=17,minutes=15)):
            problems=True
            messages.error(request,"El servicio no puede terminar luego del horario laboral (5:15pm)")
        
        #turn must not end later that midday on saturdays
        if(dateF.weekday()==5 and endHour>timedelta(hours=12)):
            problems=True
            messages.error(request,"El servicio no puede terminar luego del horario laboral de los sábados (12:00pm)")

    #turn must not be scheduled on sundays
    if(dateF.weekday()==6):
        problems=True
        messages.error(request,"El servicio no puede ser asignado en un domingo")

    if(len(actualTurns)):
        #beg hour is crashing 
        for turn in actualTurns:
            if(begHour<turn[1]):
                if(begHour>turn[0]):
                    problems=True
                    messages.error(request,"El servicio choca con el servicio de las "+delta2time(turn[0]))

                break
        #end hour is crashing 
        actualTurns.reverse()
        for turn in actualTurns:
            if(endHour>turn[0]):
                if(endHour<turn[1]):
                    problems=True
                    messages.error(request,"El servicio choca con el servicio de las "+delta2time(turn[0]))

                break
        #it is something in the middle
        for turn in actualTurns:
            if(endHour>turn[1] and begHour<turn[0]):
                problems=True
                messages.error(request,"El servicio choca con el servicio de las "+delta2time(turn[0]))

                break

    return problems

@login_required(login_url='')
def valBlocks(request, cvs, dateF, duration, startHour):
    problems = False
    dateBlocks=Block.objects.filter(cvs__name=cvs).filter(startDate=dateF).order_by('startHour')
    actualBlocks=[]
    for block in dateBlocks:
        beg=timedelta(hours=block.startHour.hour,minutes=block.startHour.minute)
        end=beg+timedelta(minutes=block.duration)
        actualBlocks.append([beg,end])
    
    begHour=timedelta(hours=startHour.hour,minutes=startHour.minute)
    endHour=begHour+timedelta(minutes=duration)
    if(len(actualBlocks)):
        #beg hour is crashing 
        for block in actualBlocks:
            if(begHour<block[1]):
                if(begHour>block[0]):
                    problems=True
                    messages.error(request,"El servicio choca con el bloqueo de las "+delta2time(block[0]))

                break
        #end hour is crashing 
        actualBlocks.reverse()
        for block in actualBlocks:
            if(endHour>block[0]):
                if(endHour<block[1]):
                    problems=True
                    messages.error(request,"El servicio choca con el bloqueo de las "+delta2time(block[0]))

                break
        #it is something in the middle
        for turn in actualBlocks:
            if(endHour>block[1] and begHour<block[0]):
                problems=True
                messages.error(request,"El servicio choca con el bloqueo de las "+delta2time(block[0]))

                break
    
    return problems

@login_required(login_url='')
def valBill(request,bill):
    problems=False
    #quantity must be an integer
    if(bill%1>0):
        problems=True
        messages.error(request,"El número de factura no puede ser fraccionaria")
    #quantity must be a positive integer
    if(bill<=0):
        problems=True
        messages.error(request,"El número de factura debe ser mayor que 0")
    #tamaño?
    return problems

@login_required(login_url='')
def valId(request,id):
    problems=False
    #quantity must be an integer
    if(id%1>0):
        problems=True
        messages.error(request,"El id del cliente no puede ser fraccionaria")
    #quantity must be a positive integer
    if(id<=0):
        problems=True
        messages.error(request,"El id del cliente no puede ser 0")

    return problems

@login_required(login_url='')
def confirm_turns(request):
    warningPage=redirectInvalidPage(request,[1,2])
    if(warningPage):
        return redirect(warningPage)

    if(request.method=='POST'):
        #validated data
        cvsName=request.POST.get('cvsName')
        typeTire=request.POST.get('typeTire')
        quantity=request.POST.get('quantity')
        rotation=request.POST.get('rotation')
        quantityRotate=request.POST.get('quantityRotate')
        duration=request.POST.get('duration')
        #data to validate
        dateF=date.fromisoformat(request.POST.get('date'))
        hour=time.fromisoformat(request.POST.get('hour'))
        flag1=valDateHour(request,cvsName,dateF,hour,int(duration),1,None,False)
        flag4 = valBlocks(request, cvsName, dateF, int(duration),hour)
        bill=int(request.POST.get('bill'))
        flag2=valBill(request,bill)
        idCustomer=int(request.POST.get('idCustomer'))
        flag3=valId(request,idCustomer)
        nameCustomer=request.POST.get('nameCustomer')
        telCustomer=request.POST.get('telCustomer')

        if(flag1 or  flag2 or flag3 or flag4):
            minDate=date.today().isoformat()
            if(date.today().weekday==6):
                maxDate=(date.today()+timedelta(days=8)).isoformat()
            else:
                maxDate=(date.today()+timedelta(days=7)).isoformat()
            deltaMax=timedelta(hours=17,minutes=15)
            deltaDuration=timedelta(minutes=int(duration))
            s = (deltaMax-deltaDuration).seconds
            hours,remainder = divmod(s, 3600)
            minutes= remainder//60
            maxTime=('{:02}:{:02}'.format(int(hours), int(minutes)))
            #debug
            print(minDate)
            print(maxDate)
            print(maxTime)
            #events' array (freetime and services) for the context
            sc_days=toSee_schedules(cvsName)
            return render(request,template_name="1-2-select_turns.html", context={'cvsName':cvsName,'sc_days':sc_days,'typeTire':typeTire,'quantity':quantity,'rotation':rotation,'quantityRotate':quantityRotate,'duration':duration,'minDate':minDate,'maxDate':maxDate,'maxTime':maxTime,'role':getRole(request)})
        else:
            bill=int(bill)
            idCustomer=int(idCustomer)
            newTurn=Turn(
                cvs = CVS.objects.get(name=cvsName),
                typeTire = typeTire,
                quantity = quantity,
                rotation = bool(rotation),
                quantityRotate = quantityRotate,
                duration = duration,
                date = dateF,
                hour = hour,
                bill = bill,
                idCustomer = idCustomer,
                nameCustomer = nameCustomer,
                telCustomer = telCustomer,
                scheduledBy = Profile.objects.get(user=request.user)
            )
            newTurn.save()
            messages.success(request,"Se ha agendado correctamente el turno en CVS: "+cvsName+" el día "+dateF.strftime("%d/%m/%Y")+" a las "+hour.strftime("%I:%M %p"))

            return redirect('see_schedules')

#---------- auxiliar methods for modify turns -------------------#

@login_required(login_url='')
def modify_schedules(request):
    warningPage=redirectInvalidPage(request,[1])
    if(warningPage):
        return redirect(warningPage)
    
    #we take the name of the admin -> cvsName
    cvsName=request.user.first_name
    #events' array (freetime and services) for the context
    sc_days=toSee_schedules(cvsName)
    #limit for the date input
    if(date.today().weekday==6):
        maxDate=(date.today()+timedelta(days=8)).isoformat()
    else:
        maxDate=(date.today()+timedelta(days=7)).isoformat()

    if(request.method=='POST'):
        dateToSearch=date.fromisoformat(request.POST.get('date'))
        turns=Turn.objects.filter(date=dateToSearch).filter(cvs__name=cvsName).order_by('hour')
        listTurns=[]
        for turn in turns:
            turnL=[]
            turnL.append(turn.id)
            beg=timedelta(hours=turn.hour.hour,minutes=turn.hour.minute)
            end=beg+timedelta(minutes=turn.duration)
            turnL.append(delta2time(beg)+" - "+delta2time(end))
            listTurns.append(turnL)
        
        if(listTurns):
            return render(request,template_name="1-modify_schedules.html",context={'cvsName':cvsName,'sc_days':sc_days,'maxDate':maxDate,'listTurns':listTurns,'date':dateToSearch.isoformat(),'role':getRole(request)})
        else:
            messages.warning(request,"El día accedido no tiene turnos asignados")

    return render(request,template_name="1-modify_schedules.html",context={'cvsName':cvsName,'sc_days':sc_days,'maxDate':maxDate,'role':getRole(request)})

@login_required(login_url='')
def modify_turn(request):
    warningPage=redirectInvalidPage(request,[1])
    if(warningPage):
        return redirect(warningPage)

    if(request.POST.get('turn')):
        #debug
        print("por turno")
        #get the turn
        turn=Turn.objects.get(id=int(request.POST.get('turn')))
    else:
        #debug
        print("por factura")
        #get the cvs
        cvsName=request.user.first_name
        #get the turn of the cvs
        turn=Turn.objects.filter(cvs__name=cvsName).filter(bill=int(request.POST.get('bill')))

        if(not len(turn)):
            messages.warning(request,"No se ha encontrado un servicio con factura: "+request.POST.get('bill')+" en este CVS")
            return redirect('modify_schedules')
        
        #get the turn from the queryset
        turn=turn.first()
    
    idTurn= turn.id
    cvsName= turn.cvs.name
    typeTire= turn.typeTire
    if(typeTire==1):
        typeTireName="Automóvil"
    elif(typeTire==2):
        typeTireName="Camioneta"
    elif(typeTire==3):
        typeTireName="Camión"
    quantity= turn.quantity
    if(turn.rotation):
        rotation="yes"
    else:
        rotation=""
    quantityRotate= turn.quantityRotate
    duration= turn.duration
    dateF= turn.date.isoformat()
    hour= turn.hour.isoformat('minutes')
    bill= turn.bill
    idCustomer= turn.idCustomer
    nameCustomer= turn.nameCustomer
    telCustomer= turn.telCustomer
    comment=turn.comment
    #450 because we have to keep the space for 
    # - the new line 
    # - the carriage return
    # - the header Mod _nameAdmin_:
    # - validation comment
    maxComment = 450-len(comment)
    #limit for the date input
    if(date.today().weekday==6):
        maxDate=(date.today()+timedelta(days=7)).isoformat()
    else:
        maxDate=(date.today()+timedelta(days=7)).isoformat()

    return render(request,template_name="1-modify_turn.html",
        context={
        'idTurn':idTurn,
        'cvsName':cvsName,
        'typeTire':typeTire,
        'typeTireName':typeTireName,
        'quantity':quantity,
        'rotation':rotation,
        'quantityRotate':quantityRotate,
        'duration':duration,
        'date':dateF,
        'hour':hour,
        'bill':bill,
        'idCustomer':idCustomer,
        'nameCustomer':nameCustomer,
        'telCustomer':telCustomer,
        'comment':comment,
        'maxComment':maxComment,
        'maxDate':maxDate,
        'role':getRole(request)})

@login_required(login_url='')
def confirm_modify(request):
    warningPage=redirectInvalidPage(request,[1])
    if(warningPage):
        return redirect(warningPage)
    
    _cvs= request.POST.get('cvsName')
    _typeTire= int(request.POST.get('typeTire'))
    _quantity= int(request.POST.get('quantity'))
    _rotation= request.POST.get('rotation')
    if(_rotation):
        _quantityRotate= int(request.POST.get('quantityRotate'))
        _rotation=True
        flag6=valQuantityRotate(request,_quantityRotate,_quantity)
    else:
        _quantityRotate=0
        _rotation=False
        flag6=False
    _duration= int(request.POST.get('duration'))
    _date= date.fromisoformat(request.POST.get('date'))
    _hour= time.fromisoformat(request.POST.get('hour'))
    _bill= int(request.POST.get('bill'))
    _idCustomer= int(request.POST.get('idCustomer'))
    _nameCustomer= request.POST.get('nameCustomer')
    _telCustomer= request.POST.get('telCustomer')

    flag1=valBill(request,_bill)
    flag3=valDuration(request,_duration)
    modTurn=Turn.objects.get(id=int(request.POST.get('idTurn')))
    flag2=valDateHour(request,_cvs,_date,_hour,_duration,0,modTurn.hour,False)
    flag4=valId(request,_idCustomer)
    flag5=valQuantity(request,_quantity)
    
    print(flag1)
    print(flag3)
    print(flag2)
    print(flag4)
    print(flag5)
    print(flag6)
    if(flag1 or flag2 or flag3 or flag4 or flag5 or flag6):
        return redirect('modify_schedules')
    else:
        modTurn.typeTire=_typeTire
        modTurn.quantity=_quantity
        modTurn.rotation=_rotation
        modTurn.quantityRotate=_quantityRotate
        modTurn.duration=_duration
        modTurn.date=_date
        modTurn.hour=_hour
        modTurn.bill=_bill
        modTurn.idCustomer=_idCustomer
        modTurn.nameCustomer=_nameCustomer
        modTurn.telCustomer=_telCustomer
        modTurn.modifiedBy=Profile.objects.get(user=request.user)
        modTurn.dateModified=datetime.now()
        modTurn.comment+="\nMod-"+Profile.objects.get(user=request.user).user.username+":\n"+request.POST.get('newComment')

        modTurn.save()
        messages.success(request,"Se actualizó correctamente el turno en CVS: "+_cvs+" para el día "+_date.strftime("%d/%m/%Y")+" a las "+_hour.strftime("%I:%M %p"))
        return redirect('see_schedules')

@login_required(login_url='')
def delete_service(request):
    warningPage=redirectInvalidPage(request,[1])
    if(warningPage):
        return redirect(warningPage)
    
    if(request.method=='POST'):
        delTurn=Turn.objects.get(id=int(request.POST.get('idTurn')))
        strTurn=str(delTurn)
        delTurn.delete()
        messages.success(request,"Se eliminó correctamente el turno "+strTurn)
        return redirect('see_schedules')
    else:
        messages.error(request,"No puedes acceder a este apartado sin haber llenado la información inicial del servicio")
        return redirect('modify_schedules')

#---------- auxiliar methods for validate service -------------------#

@login_required(login_url='')
def validate_service_provided(request):
    warningPage=redirectInvalidPage(request,[1])
    if(warningPage):
        return redirect(warningPage)
    
    #we take the name of the admin -> cvsName
    cvsName=request.user.first_name
    #events' array (freetime and services) for the context
    sc_days=toSee_schedules(cvsName)
    #limit for the date input
    if(date.today().weekday==6):
        maxDate=(date.today()+timedelta(days=8)).isoformat()
    else:
        maxDate=(date.today()+timedelta(days=7)).isoformat()

    if(request.method=='POST'):
        dateToSearch=date.fromisoformat(request.POST.get('date'))
        print(dateToSearch)
        turns=Turn.objects.filter(date=dateToSearch).filter(cvs__name=cvsName).order_by('hour')
        listTurns=[]
        for turn in turns:
            turnL=[]
            turnL.append(turn.id)
            beg=timedelta(hours=turn.hour.hour,minutes=turn.hour.minute)
            end=beg+timedelta(minutes=turn.duration)
            turnL.append(delta2time(beg)+" - "+delta2time(end))
            listTurns.append(turnL)
        
        if(listTurns):
            return render(request,template_name="1-validate_service_provided.html",context={'cvsName':cvsName,'sc_days':sc_days,'maxDate':maxDate,'listTurns':listTurns,'date':dateToSearch.isoformat(),'role':getRole(request)})
        else:
            messages.warning(request,"El día accedido no tiene turnos asignados")
    
    return render(request,template_name="1-validate_service_provided.html", context={'role':getRole(request),'cvsName':cvsName,'sc_days':sc_days,'maxDate':maxDate})

@login_required(login_url='')
def validate_turn(request):
    warningPage=redirectInvalidPage(request,[1])
    if(warningPage):
        return redirect(warningPage)

    if(request.POST.get('turn')):
        turn=Turn.objects.get(id=int(request.POST.get('turn')))
    else:
        cvsName=request.user.first_name
        #get the turn of the cvs
        turn=Turn.objects.filter(cvs__name=cvsName).filter(bill=int(request.POST.get('bill')))

        if(not len(turn)):
            messages.warning(request,"No se ha encontrado un servicio con factura: "+request.POST.get('bill')+" en este CVS")
            return redirect('validate_service_provided')
        
        #get the turn from the queryset
        turn=turn.first()
    
    idTurn= turn.id
    cvsName= turn.cvs.name
    typeTire= turn.typeTire
    if(typeTire==1):
        typeTireName="Automóvil"
    elif(typeTire==2):
        typeTireName="Camioneta"
    elif(typeTire==3):
        typeTireName="Camión"
    quantity= turn.quantity
    if(turn.rotation):
        rotation="yes"
    else:
        rotation=""
    quantityRotate= turn.quantityRotate
    duration= turn.duration
    dateF= turn.date.isoformat()
    hour= turn.hour.isoformat('minutes')
    bill= turn.bill
    idCustomer= turn.idCustomer
    nameCustomer= turn.nameCustomer
    telCustomer= turn.telCustomer
    comment=turn.comment
    done = turn.done
    #limit for the date input
    if(date.today().weekday==6):
        maxDate=(date.today()+timedelta(days=7)).isoformat()
    else:
        maxDate=(date.today()+timedelta(days=7)).isoformat()

    return render(request,template_name="1-validate_turn.html",
        context={
        'idTurn':idTurn,
        'cvsName':cvsName,
        'typeTire':typeTire,
        'typeTireName':typeTireName,
        'quantity':quantity,
        'rotation':rotation,
        'quantityRotate':quantityRotate,
        'duration':duration,
        'date':dateF,
        'hour':hour,
        'bill':bill,
        'idCustomer':idCustomer,
        'nameCustomer':nameCustomer,
        'telCustomer':telCustomer,
        'comment':comment,
        'done':done,
        'maxDate':maxDate,
        'role':getRole(request)})

@login_required(login_url='')
def confirm_validation(request):
    warningPage=redirectInvalidPage(request,[1])
    if(warningPage):
        return redirect(warningPage)
    
    _cvs= request.POST.get('cvsName')
    _comment= request.POST.get('comment')
    if(_comment):
        _comment=_comment
    else:
        _comment=""
    _newComment= request.POST.get('newComment')
    _done= request.POST.get('done')
    if(_done):
        _done=True
    else:
        _done=False
      
    _date= date.fromisoformat(request.POST.get('date'))
    _hour= time.fromisoformat(request.POST.get('hour'))

    modTurn=Turn.objects.get(id=int(request.POST.get('idTurn')))
    modTurn.done=_done
    modTurn.modifiedBy=Profile.objects.get(user=request.user)
    modTurn.dateModified=datetime.now()
    modTurn.comment=_comment + " - " + _newComment

    modTurn.save()
    messages.success(request,"Se actualizó correctamente el turno en CVS: "+_cvs+" para el día "+_date.strftime("%d/%m/%Y")+" a las "+_hour.strftime("%I:%M %p"))
    return redirect('see_schedules')